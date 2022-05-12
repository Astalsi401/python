from openpyxl import load_workbook

pwd = 'D:/Documents/work/htmlfix/snq/map'
for a in ['標章名單_中.南_129項v2', '標章名單_北.東_115項v2']:
    wb = load_workbook(f'{pwd}/{a}.xlsx')
    sheetnames = wb.get_sheet_names()
    with open(f'{pwd}/{a}_sheetname.txt', 'w', encoding='UTF-8-sig') as f:
        for b in sheetnames:
            f.write(f'{b}\n')
    print(f'{a}_sheetname.txt saved')
