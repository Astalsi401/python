import os
import logging
import pandas as pd
from csv import writer, reader
from datetime import datetime
from traceback import format_exc
from time import sleep
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
from random import uniform
from json import load

pwd = os.path.dirname(os.path.abspath(__file__)).replace('\\', '/')
chop = Options()
chop.add_extension(f'{pwd}/extensions/adblock_4.46.2_0.crx')
chop.add_extension(f'{pwd}/extensions/windscribe_3.4.3_0.crx')


def error(e):
    logging.warning(e)
    with open(f'{pwd}/.tmp/log.txt', mode='a+', encoding='utf-8') as f:
        time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        f.write(f'{time}\r\n')
        f.write(f'{str(format_exc())}\r\n')
        f.write(f'{str(e)}\r\n\r\n')
        logging.warning(f'error message saved in {pwd}/.tmp/log.json')


def alpha(alpha):
    '''
    26進位英數互換
    '''
    if type(alpha) == str:
        alpha = alpha.upper()
        assert(isinstance(alpha, str))
        return sum([(ord(n) - 64) * 26**i for i, n in enumerate(list(alpha)[::-1])])
    elif type(alpha) == int:
        assert(isinstance(alpha, int) and alpha > 0)
        num = [chr(i) for i in range(65, 91)]
        ret = []
        while alpha > 0:
            alpha, m = divmod(alpha - 1, len(num))
            ret.append(num[m])
        return ''.join(ret[::-1])


def xlsx(f, sheet, data, start):
    '''
    引入openpyxl後再使用
    list to excel
    f = openpyxl.load_workbook(path)
    sheet = sheet name
    data = 資料(list)
    start = excel開始的位置，如['A', '1']
    '''
    try:
        ws = f[sheet]
    except KeyError:
        f.create_sheet(sheet, 0)
        ws = f[sheet]
    for row in ws['A1':'H500']:
        for c in row:
            c.value = None
    start = start[0] + start[1]
    end = alpha(alpha(start[0]) + len(data[0]) - 1) + str(int(start[1]) + len(data) - 1)
    for i, r in enumerate(ws[start:end]):
        for j, c in enumerate(r):
            c.value = data[i][j]


def writeCsv(path, name, data, mode='w+', enc='utf-8-sig'):
    '''list to csv'''
    if not os.path.isdir(path):
        os.makedirs(path)
    with open(f'{path}/{name}', mode=mode, encoding=enc, newline='') as f:
        for a in data:
            writer(f).writerow(a)
    logging.info(f'{name} saved!')


def readCsv(path, name, enc='utf-8-sig'):
    '''csv to list'''
    res = []
    with open(f'{path}/{name}', mode='r', encoding=enc, newline='') as f:
        for r in [a for a in reader(f)]:
            if r != []:
                res.append(r)
    return res


def spread(arg):
    res = []
    for i in arg:
        if isinstance(i, list):
            res.extend(i)
        else:
            res.append(i)
    return res


def vpn(driver, wd, mode):
    logging.info(f'vpn {mode}')
    driver.switch_to.window(wd['windscribe'])
    w = WebDriverWait(driver, 20)
    btn = w.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-frame"]/div/div[4]/div[1]/div/div[1]/div/div[3]/button')))
    if mode == 'on':
        if btn.get_attribute('class') == 'css-1ek69w9-SimpleButton-baseStyle-SimpleButtonStyle-ButtonContainer euzhj6i0':
            btn.click()
    elif mode == 'off':
        if btn.get_attribute('class') == 'css-37am5c-SimpleButton-baseStyle-SimpleButtonStyle-ButtonContainer euzhj6i0':
            btn.click()
    driver.switch_to.window(wd['crawler'])


def vpnLogin(driver, wd):
    driver.switch_to.window(wd['windscribe'])
    vpn = 'chrome-extension://hnmpcagpplmpfojmgmnngilcnanddlhb/popup.html'
    driver.get(vpn)
    with open(f'{pwd}/.tmp/vpn.json', mode='r', encoding='utf-8') as f:
        log = load(f)
    username = log['username']
    password = log['password']
    w = WebDriverWait(driver, 20)
    w.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-frame"]/div/button[2]'))).click()
    logging.info('login windscribe')
    w.until(EC.presence_of_element_located((By.XPATH, '//*[@id="app-frame"]/div/div/form/div[1]/div[2]/input')))
    driver.find_element(By.XPATH, '//*[@id="app-frame"]/div/div/form/div[1]/div[2]/input').send_keys(username)
    w.until(EC.presence_of_element_located((By.XPATH, '//*[@id="app-frame"]/div/div/form/div[2]/div[2]/input')))
    driver.find_element(By.XPATH, '//*[@id="app-frame"]/div/div/form/div[2]/div[2]/input').send_keys(password)
    driver.find_element(By.XPATH, '//*[@id="app-frame"]/div/div/form/div[3]/button').click()
    w.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app-frame"]/div/button[2]'))).click()
    driver.switch_to.window(wd['crawler'])


def checkExists(web, path, mode='text', selector='css'):
    w = WebDriverWait(web, 20)
    try:
        if path == '':
            return ''
        elif mode == 'text':
            if selector == 'xpath':
                logging.info(f'selector: {selector}')
                logging.info(f'text: {web.find_element(By.XPATH, path).text}')
                return web.find_element(By.XPATH, path).text
            else:
                return web.find_element(By.CSS_SELECTOR, path).text
        elif mode == 'href':
            return web.find_element(By.CSS_SELECTOR, path).get_attribute('href')
    except NoSuchElementException:
        return ''
    except Exception as e:
        error(e)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return f'[{now}] error'


def getData(product, j, i, row, page=1, selector='css'):
    res = {
        'source': row['source'],
        'key': row['key'],
        'site': checkExists(product, row['site']),
        'price': checkExists(product, row['price'], selector=selector),
        'name': checkExists(product, row['name'], selector=selector),
        'event': spread([checkExists(product, e) for e in row['event']]),
        'link': checkExists(product, row['link'], 'href')
    }
    df = pd.DataFrame.from_dict([res])
    source = row['source']
    if i == 0 and j == 0 and page == 1:
        df.to_csv(f'{pwd}/csv/{source}.csv', index=False, header=True, mode='w+')
    else:
        df.to_csv(f'{pwd}/csv/{source}.csv', index=False, header=False, mode='a+')


def openSite(driver, url, limit=5):
    req = 1
    while True:
        try:
            driver.get(url)
            sleep(uniform(3, 6))
        except Exception as e:
            if req == limit:
                break
            else:
                error(e)
                logging.warning(f'retry {req}/{limit}')
                sleep(uniform(3, 6))
                continue
        break


def pgNext(driver, i, keywords, row, prodsPath, nextPath=None, mode=['css', 'css'], limit=3, count=1, page=1):
    w = WebDriverWait(driver, 20)
    while True:
        try:
            sleep(uniform(3, 6))
            logging.info(f'keywords: {i+1}/{len(keywords)}, page: {page}, {row["source"]}')
            if mode[0] == 'css':
                for j, product in enumerate(w.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, prodsPath)))):
                    getData(product, j, i, row, page)
            elif mode[0] == 'xpath':
                for j, product in enumerate(w.until(EC.presence_of_all_elements_located((By.XPATH, prodsPath)))):
                    getData(product, j, i, row, page)
        except TimeoutException:
            if count == limit:
                break
            else:
                driver.refresh()
                count += 1
                continue
        try:
            if nextPath:
                if mode[1] == 'css':
                    driver.find_element(By.CSS_SELECTOR, nextPath).click()
                elif mode[1] == 'xpath':
                    driver.find_element(By.XPATH, nextPath).click()
                page += 1
            else:
                break
        except NoSuchElementException:
            break


def momoLogin(driver):
    try:
        if driver.find_element(By.CSS_SELECTOR, '#LOGINSTATUS').get_attribute('title') == '客戶登出':
            pass
        else:
            driver.find_element(By.CSS_SELECTOR, '#LOGINSTATUS').click()
            with open(f'{pwd}/.tmp/momo.json', mode='r', encoding='utf-8') as f:
                log = load(f)
            phone = log['phone']
            password = log['password']
            sleep(5)
            driver.find_element(By.CSS_SELECTOR, '#loginForm #memId').send_keys(phone)
            driver.find_element(By.CSS_SELECTOR, '#loginForm #passwd_show').click()
            driver.find_element(By.CSS_SELECTOR, '#loginForm #passwd').send_keys(password)
            sleep(2)
            driver.find_element(By.CSS_SELECTOR, '.loginBtn input').click()
            sleep(5)
    except Exception as e:
        error(e)
        pass


def momoCoupon(driver):
    w = WebDriverWait(driver, 20)
    writeCsv(f'{pwd}/csv', 'momo折價券.csv', [['source', 'key', 'price', 'name', '折價券名稱', '折價券面額', '折價後金額', 'link']], mode='w+')
    urls = []
    for url in readCsv(f'{pwd}/csv', 'momo.csv'):
        if '折價券' in url[5]:
            urls.append(url)
    for url in urls:
        driver.get(url[6])
        sleep(uniform(3, 6))
        momoLogin(driver)
        showCoupon = w.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.showCoupon')))
        driver.execute_script('arguments[0].click();', showCoupon)
        for trs in w.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.couponList tbody tr'))):
            tds = trs.find_elements(By.CSS_SELECTOR, 'td')
            row = [url[0], url[1], url[3], url[4], tds[0].text, tds[1].text, tds[2].text, url[6]]
            writeCsv(f'{pwd}/csv', 'momo折價券.csv', [row], mode='a+')


def momo(driver, keywords):
    for i, key in enumerate(keywords):
        openSite(driver, f'https://www.momoshop.com.tw/search/searchShop.jsp?keyword={key[0]}')
        remind = driver.find_element(By.CSS_SELECTOR, '.remindBox').text
        if remind != '':
            row = ['momo', key[0], remind, '', '', '', '']
            if i == 0:
                writeCsv(f'{pwd}/csv', 'momo.csv', [row], mode='w+')
            else:
                writeCsv(f'{pwd}/csv', 'momo.csv', [row], mode='a+')
        else:
            row = {
                'source': 'momo',
                'key': key[0],
                'site': '',
                'price': '.price',
                'name': '.prdName',
                'event': ['.iconArea'],
                'link': 'a.goodsUrl[href]',
            }
            pgNext(driver, i, keywords, row, '.listArea li', "//*[contains(text(), '下一頁')]", mode=['css', 'xpath'])


def watsons(driver, keywords):
    for i, key in enumerate(keywords):
        driver.set_window_size(1080, 1000)
        openSite(driver, f'https://www.watsons.com.tw/search?text={key[0]}&pageSize=64')
        remind = driver.find_element(By.CSS_SELECTOR, '.SearchResultText.has-components.ng-star-inserted').text
        if '找不到相符結果' in remind:
            row = ['watsons', key[0], remind, '', '', '', '']
            if i == 0:
                writeCsv(f'{pwd}/csv', 'watsons.csv', [row], mode='w+')
            else:
                writeCsv(f'{pwd}/csv', 'watsons.csv', [row], mode='a+')
        else:
            row = {
                'source': 'watsons',
                'key': key[0],
                'site': '',
                'price': '.productPrice',
                'name': '.productName',
                'event': ['.productHighlight'],
                'link': '.productName a[href]',
            }
            pgNext(driver, i, keywords, row, 'e2-product-list e2-product-tile')


def google(driver, keywords):
    for i, key in enumerate(keywords):
        openSite(driver, f'https://www.google.com/search?q={key[0]}&tbm=shop')
        row = {
            'source': 'google',
            'key': key[0],
            'site': '.aULzUe.IuHnof',
            'price': '.a8Pemb.OFFNJ',
            'name': '.Xjkr3b',
            'event': ['.vEjMR'],
            'link': 'a.shntl[href]',
        }
        pgNext(driver, i, keywords, row, '.sh-pr__product-results-grid.sh-pr__product-results .sh-dgr__content')


def findprice(driver, keywords):
    for i, key in enumerate(keywords):
        openSite(driver, f'https://www.findprice.com.tw/g/{key[0]}')
        row = {
            'source': 'findprice',
            'key': key[0],
            'site': '.GoodsMname',
            'price': '.rec-price-20',
            'name': '.GoodsGname',
            'event': ['.act_div', '.discount_div'],
            'link': '.GoodsGname a.ga[href]',
        }
        pgNext(driver, i, keywords, row, '#g_div .divGoods', '#pg-next')


def cosmed(driver, keywords):
    for i, key in enumerate(keywords):
        url = f'https://shop.cosmed.com.tw/v2/Search?shopId=2131&q={key[0]}'
        driver.get(url)
        row = {
            'source': 'cosmed',
            'key': key[0],
            'site': '',
            'price': '.sc-fzXfNR.bsRQkb',
            'name': '.sc-fzXfNO.cjNSCe',
            'event': [''],
            'link': 'a[href]',
        }
        pgNext(driver, i, keywords, row, '.sc-fzXfPG.ctjckg')


def exportXlsx():
    try:
        f = load_workbook(filename=f'{pwd}/findprice.xlsx')
    except FileNotFoundError:
        f = Workbook()
    for file in ['cosmed.csv', 'findprice.csv', 'google.csv', 'momo.csv', 'momo折價券.csv', 'watsons.csv']:
        data = readCsv(f'{pwd}/csv', file)
        sheet = file.replace('.csv', '')
        xlsx(f, sheet, data, ['A', '1'])
    f.save(f'{pwd}/findprice.xlsx')


def main1(site, driver, wd):
    keywords = readCsv(f'{pwd}', 'keywords.csv')
    if site == 'google':
        vpn(driver, wd, 'off')
        google(driver, keywords)
        vpn(driver, wd, 'on')
    elif site == 'watson':
        watsons(driver, keywords)
    elif site == 'findprice':
        findprice(driver, keywords)
    elif site == 'cosmed':
        cosmed(driver, keywords)
    elif site == 'momo':
        momo(driver, keywords)
        # vpn(driver, wd, 'off')
        # momoCoupon(driver)
        # vpn(driver, wd, 'on')


def main():
    logging.basicConfig(format='%(levelname)s:%(message)s', level=logging.INFO)
    logging.info('Start Find Price')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chop)
    js = 'window.open();'
    for i in range(0, 2):
        driver.execute_script(js)
    chwnd = driver.window_handles
    wd = {
        "windscribe": driver.window_handles[len(chwnd) - 1],
        "crawler": driver.window_handles[len(chwnd) - 2]
    }
    vpnLogin(driver, wd)
    vpn(driver, wd, 'off')
    vpn(driver, wd, 'on')
    for site in readCsv(pwd, 'siteList.csv'):
        main1(site[0], driver, wd)
    driver.close()
    driver.quit()
    logging.info('Output Excel')
    exportXlsx()
    logging.info('End Find Price')


if __name__ == '__main__':
    main()