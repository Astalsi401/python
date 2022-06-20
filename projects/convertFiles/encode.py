from myfuc import readCsv
from openpyxl import load_workbook

pwd = 'D:/Documents/python/projects/crawler/findprice/csv'


def alpha(alpha):
    '''
    英文字母與數字互換
    '''
    if type(alpha) == str:
        alpha = alpha.upper()
        assert(isinstance(alpha, str))
        return sum([(ord(n) - 64) * 26**i for i, n in enumerate(list(alpha)[::-1])])
    elif type(alpha) == int:
        assert(isinstance(alpha, int) and alpha > 0)
        num = [chr(i) for i in range(65, 91)]
        ret = []
        while alpha > 0:
            alpha, m = divmod(alpha - 1, len(num))
            ret.append(num[m])
        return ''.join(ret[::-1])


def xlsx(f, sheet, data, start):
    '''
    list匯出為excel
    f = openpyxl.load_workbook(path)
    sheet = sheet name
    data = 資料(list)
    start = excel開始的位置，如['A', '1']
    '''
    ws = f[sheet]
    start = start[0] + start[1]
    end = alpha(alpha(start[0]) + len(data[0]) - 1) + str(int(start[1]) + len(data) - 1)
    for i, r in enumerate(ws[start:end]):
        for j, c in enumerate(r):
            c.value = data[i][j]


def main():
    f = load_workbook(filename=f'{pwd}/findprice.xlsx')
    for file in ['cosmed.csv', 'findprice.csv', 'google.csv', 'momo.csv', 'momo折價券.csv', 'watsons.csv']:
        data = readCsv(pwd, file)
        sheet = file.replace('.csv', '')
        xlsx(f, sheet, data, ['A', '1'])
    f.save(f'{pwd}/findprice.xlsx')


main()
