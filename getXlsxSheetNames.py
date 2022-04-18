from openpyxl import load_workbook
wb = load_workbook(
    'D:/Documents/work/2021資料庫統整/xlsx/2021國際/國際醫院主席會議 5 December 2019.xlsx')
sheetnames = wb.get_sheet_names()
with open('D:/Documents/work/2021資料庫統整/xlsx/2021國際/sheetname.txt', 'w', encoding='UTF-8') as f:
    for a in sheetnames:
        f.write(f'{a}\n')
