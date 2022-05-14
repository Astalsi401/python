import os
import requests
from csv import writer, reader
from json import dump
from datetime import date
from openpyxl import load_workbook
from time import sleep
import pandas as pd


def writeCsv(path, name, data, mode='w+'):
    if not os.path.isdir(path):
        os.makedirs(path)
    with open(f'{path}/{name}', mode=f'{mode}', encoding='utf-8-sig', newline='') as f:
        for a in data:
            writer(f).writerow(a)
    print(f'{name} saved!')


def readCsv(path, name):
    a = []
    with open(f'{path}/{name}', mode='r', encoding='utf-8-sig', newline='') as f:
        b = reader(f)
        for c in b:
            a.append(c)
    return a


def writeJson(path, json, name, mode='w+', encoding='utf-8-sig'):
    if not os.path.isdir(f'{path}/json'):
        os.makedirs(f'{path}/json')
    with open(f'{path}/{name}', mode, encoding=encoding) as f:
        dump(json, f, ensure_ascii=False)
        print(f'{path}/{name} saved!')


def convertToJson(data):
    json = []
    for a in data:
        i = 0
        b = {}
        for c in a:
            b[f'row{i}'] = c
            i += 1
        json.append(b)
    return json


def getXlsxSheets(path, name):
    wb = load_workbook(f'{path}/{name}.xlsx')
    sheetnames = wb.get_sheet_names()
    with open(f'{path}/{name}_sheetname.txt', 'w', encoding='UTF-8-sig') as f:
        for b in sheetnames:
            f.write(f'{b}\n')
    print(f'{name}_sheetname.txt saved')


def getFilesName(path, ext=None):
    if ext == None:
        return [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
    else:
        return [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f)) and f'.{ext}' in f]


def yearsCalc(yearsAgo=0):
    '''列出最近3年年份'''
    a = []
    for b in range(0, yearsAgo):
        a.append(str(date.today().year - b))
    return a
